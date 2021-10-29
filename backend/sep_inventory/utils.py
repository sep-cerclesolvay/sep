import pandas as pd

from datetime import datetime
from django.utils import timezone
from django.http import HttpResponse
from io import BytesIO
from openpyxl.utils import get_column_letter


def excel_response_from_df(df, filename, sheet_name=None, multi_index=False, worksheet_callback=None):
    if sheet_name == None:
        sheet_name = filename
    with BytesIO() as b:
        # Use the StringIO object as the filehandle.
        writer = pd.ExcelWriter(b, engine='openpyxl')
        if multi_index:
            df.to_excel(writer, sheet_name=sheet_name)
            writer.sheets[sheet_name].delete_rows(3)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]
        for i in range(2 if multi_index else 1, ws.max_column+1):
            ws.column_dimensions[get_column_letter(i)].bestFit = True
            ws.column_dimensions[get_column_letter(i)].auto_size = True

        if worksheet_callback:
            worksheet_callback(ws)

        writer.save()

        # Set up the Http response.
        response = HttpResponse(
            b.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response


def now():
    try:
        return timezone.localtime(timezone.now()).strftime('%Y-%m-%dT%H:%M:%S')

    except Exception as exp:
        print('TimeZone is not set - {}'.format(exp))
        return datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
